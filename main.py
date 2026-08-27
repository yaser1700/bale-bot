import os
import re
import json
import time
import threading
import requests

from flask import Flask, request
from openpyxl import load_workbook
from urllib.parse import unquote
from datetime import datetime


# =========================================================
# CONFIG
# =========================================================

TOKEN = os.getenv("BALE_TOKEN", "").strip()

BASE_URL = (
    f"https://tapi.bale.ai/bot{TOKEN}"
    if TOKEN
    else ""
)

PHONE = "09377700031"

WEBSITE = "https://www.tecnoyadakabbasi.ir"

ADMIN_CHAT_ID = os.getenv(
    "ADMIN_CHAT_ID",
    "570728574"
).strip()

BASE_DIR = os.path.dirname(
    os.path.abspath(__file__)
)

ORDERS_FILE = os.path.join(
    BASE_DIR,
    "orders.json"
)

app = Flask(__name__)


# =========================================================
# MEMORY
# =========================================================

CARTS = {}

USER_STATES = {}

ORDER_COUNTER = 1000

WEBHOOK_ACTIVE = False

LOCK = threading.RLock()


# =========================================================
# PDF GROUPS
# =========================================================

PDF_GROUPS = {

    "📦 سوکت عباسی":
        "سوکت عباسی",

    "🔌 کابل تکنو سبزوار":
        "کابل تکنو سبزوار",

    "⚡ وایر عباسی":
        "وایر عباسی",

    "🔩 مهره و سنسور":
        "مهره و سنسور",

    "💡 قطعات برقی خودرو":
        "قطعات برقی خودرو",

    "🧩 خارجات و پلیمریجات":
        "خارجات و پلیمریجات",

    "🔌 کابل خودرو سبزوار":
        "کابل خودرو سبزوار",

    "⚙️ شیلنگ خودرو":
        "شیلنگ خودرو",

    "⚙️ جلوبندی":
        "جلوبندی",
}


# =========================================================
# NORMALIZE
# =========================================================

def normalize(text):

    text = str(text or "")

    text = text.translate(
        str.maketrans(
            "۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩",
            "01234567890123456789"
        )
    )

    replacements = {

        "ي": "ی",
        "ى": "ی",
        "ك": "ک",
        "ۀ": "ه",
        "ة": "ه",
        "ؤ": "و",
        "إ": "ا",
        "أ": "ا",

        "\u200c": " ",
        "\u200f": "",
        "\u200e": "",

    }

    for old, new in replacements.items():

        text = text.replace(
            old,
            new
        )

    text = text.lower()

    text = re.sub(
        r"\s+",
        " ",
        text
    )

    return text.strip()


def compact(text):

    return re.sub(
        r"[^0-9a-zآ-ی]+",
        "",
        normalize(text)
    )


# =========================================================
# BALE REQUEST
# =========================================================

def bale_post(
    method,
    payload=None,
    timeout=30
):

    if not TOKEN:

        print(
            "❌ BALE_TOKEN IS EMPTY"
        )

        return None

    try:

        response = requests.post(

            f"{BASE_URL}/{method}",

            json=payload or {},

            timeout=timeout

        )

        print(

            f"BALE {method}:",
            response.status_code

        )

        if response.status_code != 200:

            print(
                response.text[:500]
            )

        return response

    except Exception as e:

        print(

            f"BALE {method} ERROR:",
            e

        )

        return None


# =========================================================
# SEND MESSAGE
# =========================================================

def send_message(
    chat_id,
    text,
    keyboard=None,
    inline=False
):

    data = {

        "chat_id": chat_id,

        "text": str(text)

    }

    if keyboard:

        if inline:

            data["reply_markup"] = {

                "inline_keyboard":
                    keyboard

            }

        else:

            data["reply_markup"] = {

                "keyboard":
                    keyboard,

                "resize_keyboard":
                    True

            }

    return bale_post(
        "sendMessage",
        data,
        30
    )


# =========================================================
# CALLBACK ANSWER
# =========================================================

def answer_callback(
    callback_id,
    text=""
):

    if not callback_id:
        return

    bale_post(

        "answerCallbackQuery",

        {

            "callback_query_id":
                callback_id,

            "text":
                text

        },

        20

    )


# =========================================================
# MAIN MENU
# =========================================================

def main_menu(chat_id):

    keyboard = [

        [
            "📞 تماس مستقیم",
            "🌐 ورود به سایت"
        ],

        [
            "🆔 آیدی من"
        ],

        [
            "📄 دریافت لیست قیمت"
        ],

        [
            "🔎 جستجوی کالا"
        ],

        [
            "🛒 سبد خرید",
            "🧾 ثبت سفارش"
        ],

        [
            "📦 سوکت عباسی",
            "🔌 کابل تکنو سبزوار"
        ],

        [
            "⚡ وایر عباسی",
            "🔩 مهره و سنسور"
        ],

        [
            "💡 قطعات برقی خودرو",
            "🧩 خارجات و پلیمریجات"
        ],

        [
            "🔌 کابل خودرو سبزوار",
            "⚙️ شیلنگ خودرو"
        ],

        [
            "⚙️ جلوبندی"
        ]

    ]

    send_message(

        chat_id,

        "📋 گزینه مورد نظر را انتخاب کنید:",

        keyboard

    )


# =========================================================
# CONTACT / WEBSITE
# =========================================================

def send_website(chat_id):

    send_message(

        chat_id,

        "🌐 برای ورود مستقیم به سایت روی دکمه زیر بزنید:",

        [[

            {
                "text":
                    "🌐 ورود مستقیم به سایت",

                "url":
                    WEBSITE
            }

        ]],

        inline=True

    )


def send_phone(chat_id):

    send_message(

        chat_id,

        "📞 برای تماس مستقیم روی دکمه زیر بزنید:",

        [[

            {
                "text":
                    "📞 تماس با ۰۹۳۷۷۷۰۰۰۳۱",

                "url":
                    "tel:+989377700031"
            }

        ]],

        inline=True

    )


# =========================================================
# PDF FIND
# =========================================================

def find_pdf(group_name):

    wanted = compact(
        group_name
    )

    try:

        files = os.listdir(
            BASE_DIR
        )

    except Exception as e:

        print(
            "PDF LIST ERROR:",
            e
        )

        return None

    for filename in files:

        if not filename.lower().endswith(
            ".pdf"
        ):
            continue

        decoded = unquote(
            filename
        )

        name = os.path.splitext(
            decoded
        )[0]

        normalized_name = compact(
            name
        )

        if (
            wanted in normalized_name
            or
            normalized_name in wanted
        ):

            return os.path.join(
                BASE_DIR,
                filename
            )

    return None


# =========================================================
# SEND PDF
# =========================================================

def send_pdf(
    chat_id,
    pdf_path,
    caption
):

    try:

        with open(
            pdf_path,
            "rb"
        ) as file:

            response = requests.post(

                f"{BASE_URL}/sendDocument",

                data={

                    "chat_id":
                        str(chat_id),

                    "caption":
                        caption

                },

                files={

                    "document": (

                        os.path.basename(
                            pdf_path
                        ),

                        file,

                        "application/pdf"

                    )

                },

                timeout=180

            )

        return response

    except Exception as e:

        print(
            "PDF ERROR:",
            e
        )

        return None


# =========================================================
# PRICE
# =========================================================

def format_price(value):

    if value is None:
        return ""

    if isinstance(
        value,
        int
    ):

        return f"{value:,}"

    if isinstance(
        value,
        float
    ):

        if value.is_integer():

            return f"{int(value):,}"

        return str(value)

    text = str(
        value
    ).strip()

    text = text.replace(
        ",",
        ""
    )

    text = text.replace(
        "٬",
        ""
    )

    try:

        number = float(
            text
        )

        if number.is_integer():

            return f"{int(number):,}"

    except Exception:
        pass

    return text


def price_number(value):

    try:

        text = str(
            value or ""
        )

        text = text.replace(
            ",",
            ""
        )

        text = text.replace(
            "٬",
            ""
        )

        text = normalize(
            text
        )

        return int(
            float(text)
        )

    except Exception:

        return 0


# =========================================================
# EXCEL COLUMNS
# =========================================================

def get_columns(row):

    columns = {}

    for index, value in enumerate(
        row
    ):

        if value is None:
            continue

        name = normalize(
            value
        )

        if "گروه" in name:

            columns["group"] = index

        elif (

            "کد کالا" in name

            or name == "کد"

            or "شناسه" in name

        ):

            columns["code"] = index

        elif (

            "نام کالا" in name

            or name == "نام"

        ):

            columns["name"] = index

        elif "قیمت" in name:

            columns["price"] = index

        elif (

            "توضیحات" in name

            or "توضیح" in name

        ):

            columns["description"] = index

    return columns


# =========================================================
# SEARCH MATCH
# =========================================================

def search_matches(
    query,
    code,
    name,
    group,
    description
):

    q = normalize(
        query
    )

    qc = compact(
        query
    )

    if not q:
        return False

    full_text = " ".join([

        normalize(code),

        normalize(name),

        normalize(group),

        normalize(description)

    ])

    full_compact = compact(
        full_text
    )

    if q in full_text:
        return True

    if (
        qc
        and
        qc in full_compact
    ):

        return True

    words = [

        word

        for word in q.split()

        if len(word) >= 2

    ]

    if not words:
        return False

    return all(

        word in full_text

        for word in words

    )


# =========================================================
# SEARCH EXCEL
# =========================================================

def search_excel(query):

    try:

        excel_files = [

            filename

            for filename in os.listdir(
                BASE_DIR
            )

            if (

                filename.lower().endswith(
                    ".xlsx"
                )

                and

                not filename.startswith(
                    "~$"
                )

            )

        ]

    except Exception as e:

        print(
            "EXCEL LIST ERROR:",
            e
        )

        return []

    results = []

    seen = set()

    print(
        "SEARCH:",
        query
    )

    for filename in excel_files:

        path = os.path.join(

            BASE_DIR,

            filename

        )

        try:

            workbook = load_workbook(

                path,

                read_only=True,

                data_only=True

            )

            for sheet in workbook.worksheets:

                header_row = None

                columns = None

                for row_number, row in enumerate(

                    sheet.iter_rows(
                        values_only=True
                    ),

                    start=1

                ):

                    found = get_columns(
                        row
                    )

                    if (

                        "code" in found

                        and

                        "name" in found

                        and

                        "price" in found

                    ):

                        header_row = row_number

                        columns = found

                        break

                if not columns:
                    continue

                for row in sheet.iter_rows(

                    min_row=header_row + 1,

                    values_only=True

                ):

                    if not row:
                        continue

                    def get_value(key):

                        index = columns.get(
                            key
                        )

                        if index is None:
                            return ""

                        if index >= len(row):
                            return ""

                        return str(
                            row[index] or ""
                        ).strip()

                    group = get_value(
                        "group"
                    )

                    code = get_value(
                        "code"
                    )

                    name = get_value(
                        "name"
                    )

                    description = get_value(
                        "description"
                    )

                    price = ""

                    price_index = columns.get(
                        "price"
                    )

                    if (

                        price_index is not None

                        and

                        price_index < len(row)

                    ):

                        price = format_price(

                            row[
                                price_index
                            ]

                        )

                    if not code and not name:
                        continue

                    if not search_matches(

                        query,

                        code,

                        name,

                        group,

                        description

                    ):

                        continue

                    key = (

                        normalize(code),

                        normalize(name),

                        normalize(group),

                        price

                    )

                    if key in seen:
                        continue

                    seen.add(
                        key
                    )

                    results.append({

                        "code":
                            code,

                        "name":
                            name,

                        "price":
                            price,

                        "group":
                            group,

                        "description":
                            description

                    })

            workbook.close()

        except Exception as e:

            print(

                "EXCEL ERROR:",

                filename,

                e

            )

    print(
        "RESULT COUNT:",
        len(results)
    )

    return results


# =========================================================
# SEARCH RESULTS
# =========================================================

def send_search_results(
    chat_id,
    results
):

    if not results:

        send_message(

            chat_id,

            "❌ کالایی با این کد یا نام پیدا نشد."

        )

        return

    with LOCK:

        USER_STATES[chat_id] = {

            "state":
                "search_results",

            "results":
                results

        }

    send_message(

        chat_id,

        f"🔎 {len(results)} کالا پیدا شد:\n\n"
        "برای هر کالا، دکمه ➕ همان کالا را بزنید."

    )

    for index, item in enumerate(
        results
    ):

        message = (

            f"📦 کد کالا: "
            f"{item['code']}\n"

            f"📝 نام کالا: "
            f"{item['name']}\n"

            f"💰 قیمت: "
            f"{item['price']} ریال\n"

            f"📁 گروه: "
            f"{item['group']}"

        )

        if item["description"]:

            message += (

                "\nℹ️ توضیحات: "

                +
                item["description"]

            )

        callback_data = (

            f"ADD:{chat_id}:{index}"

        )

        keyboard = [[

            {

                "text":
                    "➕ افزودن به سبد خرید",

                "callback_data":
                    callback_data

            }

        ]]

        send_message(

            chat_id,

            message,

            keyboard,

            inline=True

        )


# =========================================================
# CART
# =========================================================

def add_to_cart(
    chat_id,
    item,
    quantity
):

    with LOCK:

        CARTS.setdefault(
            chat_id,
            []
        )

        cart = CARTS[chat_id]

        for cart_item in cart:

            if (

                cart_item["code"]

                ==

                item["code"]

            ):

                cart_item["quantity"] += quantity

                return

        cart.append({

            "code":
                item["code"],

            "name":
                item["name"],

            "price":
                price_number(
                    item["price"]
                ),

            "quantity":
                quantity

        })


def show_cart(chat_id):

    with LOCK:

        cart = list(
            CARTS.get(
                chat_id,
                []
            )
        )

    if not cart:

        send_message(

            chat_id,

            "🛒 سبد خرید شما خالی است."

        )

        return

    message = (
        "🛒 سبد خرید شما:\n\n"
    )

    total = 0

    for index, item in enumerate(

        cart,

        start=1

    ):

        item_total = (

            item["price"]

            *

            item["quantity"]

        )

        total += item_total

        message += (

            f"{index}. "
            f"{item['name']}\n"

            f"🔢 کد: "
            f"{item['code']}\n"

            f"📦 تعداد: "
            f"{item['quantity']}\n"

            f"💰 قیمت واحد: "
            f"{item['price']:,} ریال\n"

            f"💵 مبلغ: "
            f"{item_total:,} ریال\n"

            "──────────────\n"

        )

    message += (

        f"\n💰 جمع کل: "
        f"{total:,} ریال"

    )

    keyboard = [

        ["🧾 ثبت سفارش"],

        ["🔎 جستجوی کالا"],

        ["🗑️ خالی کردن سبد"],

        ["🔙 منوی اصلی"]

    ]

    send_message(

        chat_id,

        message,

        keyboard

    )


def clear_cart(chat_id):

    with LOCK:

        CARTS[chat_id] = []

        USER_STATES.pop(
            chat_id,
            None
        )

    send_message(

        chat_id,

        "🗑️ سبد خرید خالی شد."

    )

    main_menu(
        chat_id
    )


# =========================================================
# ORDERS FILE
# =========================================================

def load_orders():

    global ORDER_COUNTER

    try:

        if not os.path.exists(
            ORDERS_FILE
        ):

            return []

        with open(

            ORDERS_FILE,

            "r",

            encoding="utf-8"

        ) as file:

            orders = json.load(
                file
            )

        if orders:

            numbers = [

                int(
                    order.get(
                        "order_number",
                        1000
                    )
                )

                for order in orders

                if str(
                    order.get(
                        "order_number",
                        ""
                    )
                ).isdigit()

            ]

            if numbers:

                ORDER_COUNTER = max(
                    numbers
                )

        return orders

    except Exception as e:

        print(
            "LOAD ORDERS ERROR:",
            e
        )

        return []


def save_order(order):

    try:

        orders = load_orders()

        orders.append(
            order
        )

        temp_file = (

            ORDERS_FILE
            +
            ".tmp"

        )

        with open(

            temp_file,

            "w",

            encoding="utf-8"

        ) as file:

            json.dump(

                orders,

                file,

                ensure_ascii=False,

                indent=2

            )

        os.replace(

            temp_file,

            ORDERS_FILE

        )

        print(

            "ORDER SAVED:",

            order["order_number"]

        )

        return True

    except Exception as e:

        print(
            "SAVE ORDER ERROR:",
            e
        )

        return False


# =========================================================
# ADMIN ORDER
# =========================================================

def send_order_to_admin(
    order
):

    if not ADMIN_CHAT_ID:

        print(
            "❌ ADMIN_CHAT_ID NOT SET"
        )

        return False

    message = (

        "🚨 سفارش جدید\n\n"

        f"🔢 شماره سفارش: "
        f"{order['order_number']}\n"

        f"📅 تاریخ: "
        f"{order['date']}\n\n"

        f"👤 مشتری: "
        f"{order['name']}\n"

        f"📞 تماس: "
        f"{order['phone']}\n"

        f"🆔 Chat ID: "
        f"{order['chat_id']}\n\n"

        "📦 کالاها:\n"

    )

    for item in order["items"]:

        message += (

            f"\n• {item['name']}\n"

            f"کد: {item['code']}\n"

            f"تعداد: {item['quantity']}\n"

            f"قیمت واحد: "
            f"{item['price']:,} ریال\n"

            f"مبلغ: "
            f"{item['total']:,} ریال\n"

        )

    message += (

        "\n────────────────\n"

        f"💰 مبلغ کل: "
        f"{order['total']:,} ریال"

    )

    response = send_message(

        ADMIN_CHAT_ID,

        message

    )

    try:

        return bool(

            response

            and

            response.json().get(
                "ok"
            )

        )

    except Exception:

        return False


# =========================================================
# START ORDER
# =========================================================

def start_order(chat_id):

    with LOCK:

        cart = list(
            CARTS.get(
                chat_id,
                []
            )
        )

    if not cart:

        send_message(

            chat_id,

            "🛒 سبد خرید شما خالی است.\n"
            "ابتدا کالا اضافه کنید."

        )

        return

    with LOCK:

        USER_STATES[chat_id] = {

            "state":
                "order_name"

        }

    send_message(

        chat_id,

        "🧾 ثبت سفارش\n\n"
        "لطفاً نام و نام خانوادگی خود را وارد کنید:"

    )


# =========================================================
# FINISH ORDER
# =========================================================

def finish_order(
    chat_id,
    name,
    phone
):

    with LOCK:

        cart = list(
            CARTS.get(
                chat_id,
                []
            )
        )

    if not cart:

        send_message(

            chat_id,

            "❌ سبد خرید خالی است."

        )

        return

    global ORDER_COUNTER

    with LOCK:

        ORDER_COUNTER += 1

        order_number = ORDER_COUNTER

    items = []

    total = 0

    for item in cart:

        item_total = (

            item["price"]

            *

            item["quantity"]

        )

        total += item_total

        items.append({

            "code":
                item["code"],

            "name":
                item["name"],

            "price":
                item["price"],

            "quantity":
                item["quantity"],

            "total":
                item_total

        })

    order = {

        "order_number":
            order_number,

        "date":
            datetime.now().strftime(
                "%Y-%m-%d %H:%M:%S"
            ),

        "chat_id":
            chat_id,

        "name":
            name,

        "phone":
            phone,

        "items":
            items,

        "total":
            total

    }

    saved = save_order(
        order
    )

    admin_sent = send_order_to_admin(
        order
    )

    message = (

        "✅ سفارش شما با موفقیت ثبت شد.\n\n"

        f"🔢 شماره سفارش: "
        f"{order_number}\n"

        f"👤 نام: "
        f"{name}\n"

        f"📞 تماس: "
        f"{phone}\n\n"

        "📦 اقلام سفارش:\n"

    )

    for item in items:

        message += (

            f"\n• {item['name']}\n"

            f"کد: {item['code']}\n"

            f"تعداد: {item['quantity']}\n"

            f"مبلغ: "
            f"{item['total']:,} ریال\n"

        )

    message += (

        "\n────────────────\n"

        f"💰 مبلغ کل: "
        f"{total:,} ریال\n\n"

        "📞 جهت پیگیری:\n"
        "۰۹۳۷۷۷۰۰۰۳۱"

    )

    send_message(

        chat_id,

        message

    )

    if not admin_sent:

        print(
            "⚠️ ADMIN ORDER SEND FAILED"
        )

    if not saved:

        print(
            "⚠️ ORDER FILE SAVE FAILED"
        )

    with LOCK:

        CARTS[chat_id] = []

        USER_STATES.pop(
            chat_id,
            None
        )

    main_menu(
        chat_id
    )


# =========================================================
# CALLBACK
# =========================================================

def process_callback(
    callback
):

    callback_id = callback.get(
        "id"
    )

    data = str(
        callback.get(
            "data"
        )
        or ""
    )

    message = (
        callback.get(
            "message"
        )
        or {}
    )

    chat = (
        message.get(
            "chat"
        )
        or {}
    )

    chat_id = chat.get(
        "id"
    )

    if not chat_id:

        answer_callback(
            callback_id,
            "خطا"
        )

        return

    if not data.startswith(
        "ADD:"
    ):

        answer_callback(
            callback_id
        )

        return

    parts = data.split(
        ":"
    )

    if len(parts) != 3:

        answer_callback(

            callback_id,

            "خطا در انتخاب کالا"

        )

        return

    try:

        result_chat_id = int(
            parts[1]
        )

        index = int(
            parts[2]
        )

    except Exception:

        answer_callback(

            callback_id,

            "خطا در انتخاب کالا"

        )

        return

    if result_chat_id != chat_id:

        answer_callback(

            callback_id,

            "این دکمه برای کاربر دیگری است."

        )

        return

    with LOCK:

        state_data = USER_STATES.get(
            chat_id
        )

    if not state_data:

        answer_callback(

            callback_id,

            "نتایج جستجو منقضی شده است."

        )

        return

    results = state_data.get(
        "results",
        []
    )

    if (

        index < 0

        or

        index >= len(results)

    ):

        answer_callback(

            callback_id,

            "کالا پیدا نشد."

        )

        return

    item = results[index]

    with LOCK:

        USER_STATES[chat_id] = {

            "state":
                "quantity",

            "item":
                item,

            "results":
                results

        }

    answer_callback(

        callback_id,

        "کالا انتخاب شد ✅"

    )

    send_message(

        chat_id,

        f"📦 {item['name']}\n\n"

        f"🔢 کد کالا: "
        f"{item['code']}\n"

        f"💰 قیمت: "
        f"{item['price']} ریال\n\n"

        "🔢 تعداد مورد نظر را وارد کنید:"

    )


# =========================================================
# MESSAGE PROCESS
# =========================================================

def process_message(
    message
):

    chat = (
        message.get(
            "chat"
        )
        or {}
    )

    chat_id = chat.get(
        "id"
    )

    text = str(
        message.get(
            "text"
        )
        or ""
    ).strip()

    if not chat_id:
        return

    print(

        "USER MESSAGE:",

        chat_id,

        text

    )

    # =====================================================
    # START
    # =====================================================

    if text == "/start":

        with LOCK:

            CARTS.setdefault(
                chat_id,
                []
            )

            USER_STATES.pop(
                chat_id,
                None
            )

        send_message(

            chat_id,

            "سلام 👋\n\n"
            "به ربات تولیدی و بازرگانی عباسی خوش آمدید."

        )

        main_menu(
            chat_id
        )

        return

    # =====================================================
    # MY ID
    # =====================================================

    if text in (

        "🆔 آیدی من",

        "/myid"

    ):

        send_message(

            chat_id,

            "🆔 آیدی عددی شما:\n\n"
            f"{chat_id}\n\n"
            "برای تنظیم دریافت سفارش‌ها:\n"
            "ADMIN_CHAT_ID"

        )

        return

    # =====================================================
    # WEBSITE
    # =====================================================

    if text == "🌐 ورود به سایت":

        send_website(
            chat_id
        )

        return

    # =====================================================
    # PHONE
    # =====================================================

    if text == "📞 تماس مستقیم":

        send_phone(
            chat_id
        )

        return

    # =====================================================
    # MAIN MENU
    # =====================================================

    if text in (

        "🔙 منوی اصلی",

        "منوی اصلی"

    ):

        with LOCK:

            USER_STATES.pop(
                chat_id,
                None
            )

        main_menu(
            chat_id
        )

        return

    # =====================================================
    # SEARCH BUTTON
    # =====================================================

    if text == "🔎 جستجوی کالا":

        with LOCK:

            USER_STATES[chat_id] = {

                "state":
                    "search"

            }

        send_message(

            chat_id,

            "🔎 کد یا نام کالا را ارسال کنید.\n\n"
            "مثال:\n"
            "100158\n\n"
            "یا:\n"
            "کابل دنا"

        )

        return

    # =====================================================
    # CART
    # =====================================================

    if text == "🛒 سبد خرید":

        with LOCK:

            USER_STATES.pop(
                chat_id,
                None
            )

        show_cart(
            chat_id
        )

        return

    # =====================================================
    # CLEAR CART
    # =====================================================

    if text == "🗑️ خالی کردن سبد":

        clear_cart(
            chat_id
        )

        return

    # =====================================================
    # ORDER
    # =====================================================

    if text == "🧾 ثبت سفارش":

        start_order(
            chat_id
        )

        return

    # =====================================================
    # STATE
    # =====================================================

    with LOCK:

        state_data = USER_STATES.get(
            chat_id
        )

    if state_data:

        state = state_data.get(
            "state"
        )

        # -------------------------------------------------
        # NAME
        # -------------------------------------------------

        if state == "order_name":

            if not text:

                send_message(

                    chat_id,

                    "❌ نام را وارد کنید."

                )

                return

            with LOCK:

                state_data["name"] = text

                state_data["state"] = (
                    "order_phone"
                )

            send_message(

                chat_id,

                "📞 لطفاً شماره تماس خود را وارد کنید:"

            )

            return

        # -------------------------------------------------
        # PHONE
        # -------------------------------------------------

        if state == "order_phone":

            phone = normalize(
                text
            )

            phone = re.sub(
                r"\D",
                "",
                phone
            )

            if len(phone) < 10:

                send_message(

                    chat_id,

                    "❌ شماره تماس صحیح نیست.\n"
                    "لطفاً دوباره وارد کنید:"

                )

                return

            finish_order(

                chat_id,

                state_data["name"],

                phone

            )

            return

        # -------------------------------------------------
        # QUANTITY
        # -------------------------------------------------

        if state == "quantity":

            number_text = normalize(
                text
            )

            if not number_text.isdigit():

                send_message(

                    chat_id,

                    "❌ تعداد باید عدد باشد.\n"
                    "مثلاً: 2"

                )

                return

            quantity = int(
                number_text
            )

            if quantity <= 0:

                send_message(

                    chat_id,

                    "❌ تعداد باید بیشتر از صفر باشد."

                )

                return

            item = state_data.get(
                "item"
            )

            results = state_data.get(
                "results",
                []
            )

            if not item:

                send_message(

                    chat_id,

                    "❌ اطلاعات کالا پیدا نشد."

                )

                return

            add_to_cart(

                chat_id,

                item,

                quantity

            )

            # مهم:
            # بعد از اضافه شدن کالا،
            # نتایج جستجو حفظ می‌شوند
            # تا کالاهای دیگر هم اضافه شوند.

            with LOCK:

                USER_STATES[chat_id] = {

                    "state":
                        "search_results",

                    "results":
                        results

                }

            send_message(

                chat_id,

                f"✅ {item['name']}\n"
                f"📦 تعداد {quantity} عدد به سبد اضافه شد.\n\n"
                "می‌توانید کالاهای دیگری هم به سبد اضافه کنید."

            )

            send_message(

                chat_id,

                "🛒 برای مشاهده سبد:",

                [[
                    "🛒 سبد خرید"
                ]]

            )

            return

    # =====================================================
    # PDF
    # =====================================================

    if text in PDF_GROUPS:

        with LOCK:

            USER_STATES.pop(
                chat_id,
                None
            )

        group = PDF_GROUPS[text]

        send_message(

            chat_id,

            "⏳ در حال آماده‌سازی فایل PDF..."

        )

        pdf_path = find_pdf(
            group
        )

        if not pdf_path:

            send_message(

                chat_id,

                "❌ فایل PDF این گروه پیدا نشد."

            )

            return

        response = send_pdf(

            chat_id,

            pdf_path,

            f"📄 لیست قیمت {group}"

        )

        try:

            if (

                response is not None

                and

                response.json().get(
                    "ok"
                )

            ):

                return

        except Exception:
            pass

        send_message(

            chat_id,

            "❌ ارسال فایل PDF انجام نشد."

        )

        return

    # =====================================================
    # GENERAL SEARCH
    # =====================================================

    if text:

        send_message(

            chat_id,

            "🔎 در حال جستجوی کالا..."

        )

        results = search_excel(
            text
        )

        send_search_results(

            chat_id,

            results

        )

        return


# =========================================================
# UPDATE
# =========================================================

def process_update(
    update
):

    try:

        if update.get(
            "callback_query"
        ):

            process_callback(

                update[
                    "callback_query"
                ]

            )

        elif update.get(
            "message"
        ):

            process_message(

                update[
                    "message"
                ]

            )

    except Exception as e:

        print(

            "UPDATE ERROR:",

            repr(e)

        )


# =========================================================
# WEBHOOK
# =========================================================

@app.route(
    "/webhook",
    methods=["POST"]
)
def webhook():

    try:

        update = request.get_json(
            silent=True
        )

        if update:

            # پاسخ سریع به Bale
            # پردازش در Thread جدا
            threading.Thread(

                target=process_update,

                args=(update,),

                daemon=True

            ).start()

        return {
            "ok": True
        }

    except Exception as e:

        print(

            "WEBHOOK ERROR:",

            repr(e)

        )

        return {
            "ok": True
        }


# =========================================================
# HEALTH
# =========================================================

@app.route("/")
def home():

    return (
        "Bale Bot is running"
    )


@app.route("/health")
def health():

    return {

        "status":
            "ok",

        "webhook":
            WEBHOOK_ACTIVE,

        "admin_configured":
            bool(
                ADMIN_CHAT_ID
            )

    }


# =========================================================
# GET WEBHOOK INFO
# =========================================================

def get_webhook_info():

    response = bale_post(

        "getWebhookInfo",

        {},

        20

    )

    if response is None:
        return None

    try:

        return response.json()

    except Exception:

        return None


# =========================================================
# SET WEBHOOK
# =========================================================

def setup_webhook():

    global WEBHOOK_ACTIVE

    if not TOKEN:

        print(
            "❌ TOKEN NOT FOUND"
        )

        return False

    render_url = os.getenv(
        "RENDER_EXTERNAL_URL",
        ""
    ).strip()

    custom_url = os.getenv(
        "WEBHOOK_URL",
        ""
    ).strip()

    if custom_url:

        webhook_url = (

            custom_url.rstrip("/")
            +
            "/webhook"

        )

    elif render_url:

        webhook_url = (

            render_url.rstrip("/")
            +
            "/webhook"

        )

    else:

        print(
            "❌ RENDER_EXTERNAL_URL NOT FOUND"
        )

        return False

    print(
        "WEBHOOK URL:",
        webhook_url
    )

    for attempt in range(
        1,
        6
    ):

        try:

            response = bale_post(

                "setWebhook",

                {

                    "url":
                        webhook_url

                },

                30

            )

            if response:

                try:

                    data = response.json()

                    print(
                        "SET WEBHOOK RESULT:",
                        data
                    )

                    if data.get(
                        "ok"
                    ):

                        WEBHOOK_ACTIVE = True

                        print(
                            "✅ WEBHOOK ACTIVE"
                        )

                        return True

                except Exception:
                    pass

        except Exception as e:

            print(
                "WEBHOOK ATTEMPT ERROR:",
                e
            )

        time.sleep(
            3
        )

    print(
        "❌ WEBHOOK COULD NOT BE SET"
    )

    return False


# =========================================================
# WEBHOOK WATCHER
# =========================================================

def webhook_watcher():

    global WEBHOOK_ACTIVE

    while True:

        try:

            time.sleep(
                300
            )

            info = get_webhook_info()

            if not info:

                continue

            result = info.get(
                "result",
                {}
            )

            current_url = str(
                result.get(
                    "url",
                    ""
                )
            )

            last_error = result.get(
                "last_error_message"
            )

            if last_error:

                print(

                    "⚠️ WEBHOOK LAST ERROR:",

                    last_error

                )

            if not current_url:

                print(
                    "⚠️ WEBHOOK MISSING"
                )

                setup_webhook()

                continue

            if not current_url.endswith(
                "/webhook"
            ):

                print(
                    "⚠️ WRONG WEBHOOK URL"
                )

                setup_webhook()

                continue

            WEBHOOK_ACTIVE = True

        except Exception as e:

            print(

                "WEBHOOK WATCHER ERROR:",

                repr(e)

            )


# =========================================================
# START
# =========================================================

if __name__ == "__main__":

    print(
        "========================================"
    )

    print(
        "BALE BOT FINAL VERSION"
    )

    print(
        "========================================"
    )

    print(
        "SEARCH: ON"
    )

    print(
        "CART: ON"
    )

    print(
        "MULTI PRODUCT CART: ON"
    )

    print(
        "ORDER: ON"
    )

    print(
        "ADMIN ID:",
        ADMIN_CHAT_ID
    )

    print(
        "PRICE UNIT: RIAL"
    )

    print(
        "WEBHOOK MODE: ON"
    )

    print(
        "========================================"
    )

    load_orders()

    setup_webhook()

    threading.Thread(

        target=webhook_watcher,

        daemon=True

    ).start()

    port = int(

        os.environ.get(
            "PORT",
            "10000"
        )

    )

    app.run(

        host="0.0.0.0",

        port=port,

        threaded=True

    )
